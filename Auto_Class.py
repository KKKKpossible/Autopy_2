from openpyxl import Workbook
from openpyxl import load_workbook
import tkinter
import pyautogui as pg  # message box library
from tkinter import filedialog as fd  # import file dialog
import os
from functools import partial
import Instrument
import time
import numpy as np
import MyCal


class ExcelVariableDeclare:
    def __init__(self, path):
        self.load_clear = True
        self.path = path
        self.l_wb = ""
        self.l_ws_list = []
        self.w_wb = ""
        self.w_ws_list = []
        self.column_name = []
        self.freq_var = []
        self.input_offset_var = []
        self.output_offset_var = []
        self.power_voltage_var = ""
        self.power_current_var = ""
        self.power_cur_var = ""
        self.power_m_probe = ""
        self.aging_var = ""
        self.aging_left_var = ""
        self.aging_done_var = False
        self.reference_done = False
        self.source_com_var = ""
        self.power_com_var = ""
        self.power_meter_com_var = ""
        self.spectrum_com_var = ""
        self.network_com_var = ""

        self.multiple_freq = 0  # Hz, KHz, MHz, GHz, THz
        self.multiple_aging = 0  # sec, min, hour
        self.source_comm_option = ""  # GPIB, USB, Serial
        self.power_comm_option = ""  # GPIB, USB, Serial
        self.power_meter_comm_option = ""  # GPIB, USB, Serial

        self.multiple_select_freq = 0  # Hz, KHz, MHz, GHz, THz
        self.select_freq_var = []  # select freq var
        self.atr_start_input_var = ""
        self.p_sat_input_var = ""
        self.overdrive_input_var = ""
        self.pout_var = ""
        self.output_offset_ref_power = ""

        self.freq_cell_name = ""  # default Frequency
        self.input_cell_name = ""  # default Input offset
        self.output_cell_name = ""  # default Output offset
        self.power_voltage_cell_name = ""  # default Power Voltage
        self.power_current_cell_name = ""  # default Power Voltage
        self.power_meter_probe_cell_name = ""  # default Power Voltage
        self.aging_cell_name = ""  # default Aging time
        self.source_comm_cell_name = ""  # default Source generate
        self.power_comm_cell_name = ""  # default Power supply
        self.power_meter_comm_cell_name = ""  # default Power meter
        self.spectrum_comm_cell_name = ""  # default Power meter
        self.network_comm_cell_name = ""  # default Power meter
        self.select_freq_cell_name = ""  # default Select Freq
        self.atr_start_input_dbm_name = ""  # default ATR start input dBm
        self.p_sat_input_name = ""  # default P_sat input
        self.output_off_ref_name = ""  # default P_sat input
        self.overdrive_input_name = ""  # default Overdrive input
        self.pout_name = ""  # default Pout

        self.source_baud_rate = ""  # default Source generate
        self.power_baud_rate = ""  # default Power supply
        self.power_meter_baud_rate = ""  # default Power meter


class CommonDialogVar:
    def __init__(self):
        # excel instance call
        self.excel = Excel()
        # dialog set
        self.g_root = tkinter.Tk()
        self.g_root.protocol("WM_DELETE_WINDOW", self.root_close)
        self.g_root.withdraw()
        self.excel_path = ""

        self.remote_p_meter_id = 0
        self.remote_power_id = 1
        self.remote_network_id = 2
        self.remote_source_id = 3
        self.remote_spectrum_id = 4
        self.remote_all_id = 5

        self.atr_power_id = 6
        self.atr_network_id = 7
        self.atr_spectrum_id = 8
        self.atr_offset_id = 9

        self.atr_start_button_id = 0
        self.atr_stop_button_id = 1
        self.atr_display_button_id = 2
        self.atr_save_button_id = 3
        self.atr_measure_input_offset_id = 4
        self.atr_measure_output_offset_id = 5
        self.atr_save_all_offset_id = 6

        self.remote_frame = ""
        self.atr_frame = ""

        self.remote_start_x = 30
        self.remote_start_y = 10
        self.remote_button_x = 10
        self.remote_y_gap = 10

        self.atr_start_x = 25
        self.atr_start_y = 10
        self.atr_button_x = 10
        self.atr_y_gap = 10

        self.main_butt_width = 10
        self.main_butt_height = 1

        # atr entry values
        self.atr_freq = tkinter.StringVar()
        self.atr_input_offset = tkinter.StringVar()
        self.atr_output_offset = tkinter.StringVar()
        self.atr_power_voltage = tkinter.StringVar()
        self.atr_power_current = tkinter.StringVar()
        self.atr_aging_time = tkinter.StringVar()
        self.atr_rf_input = tkinter.StringVar()
        self.atr_rf_output = tkinter.StringVar()
        self.atr_current = tkinter.StringVar()
        self.atr_aging_left = tkinter.StringVar()
        self.atr_system_option = False
        self.atr_sys_fwd = tkinter.StringVar()
        self.atr_sys_freq = tkinter.StringVar()
        self.atr_sys_temp = tkinter.StringVar()
        self.atr_sys_cmd = tkinter.StringVar()

        self.atr_source_com = tkinter.StringVar()
        self.atr_power_com = tkinter.StringVar()
        self.atr_power_meter_com = tkinter.StringVar()

        self.atr_rf_var = []
        self.atr_current_var = []

        self.atr_state = ""
        self.atr_p1_ref_input = ""
        self.atr_input_input = ""
        self.atr_p_sat_input = ""
        self.atr_index = 0

        self.atr_p1_var = []
        self.atr_input_var = []
        self.atr_input_curr_var = []
        self.atr_p_sat_var = []
        self.atr_p_sat_curr_var = []
        self.atr_overdrive_var = []
        self.atr_overdrive_curr_var = []

        self.mes_input_offset = []
        self.mes_output_offset = []

        self.inst_source = ""
        self.inst_power = ""
        self.inst_power_meter = ""

        self.atr_stop = False

        self.atr_seq = 0
        self.atr_compare_0 = 0.0
        self.atr_compare_1 = 0.0
        self.adder = 0.0

    def root_close(self):
        self.g_root.withdraw()
        self.g_root.quit()


class ExcelCommonVar:
    def __init__(self):
        self.excel_data_name_row = 1
        self.excel_data_unit_row = 2
        self.excel_data_start_row = 3
        self.excel_freq_unit = ["Hz", "KHz", "MHz", "GHz", "THz"]
        self.excel_input_unit = ["dB"]
        self.excel_output_unit = ["dB"]
        self.excel_power_unit = ["V"]
        self.excel_current_unit = ["A"]
        self.excel_power_meter_probe_unit = ["Channel"]
        self.excel_aging_unit = ["sec", "min", "hour"]
        self.excel_com_unit = ["GPIB", "USB", "Serial"]
        self.excel_dbm_unit = ["dBm"]
        self.rf_power = []

        self.excel_freq_column = "A"
        self.excel_input_column = "B"
        self.excel_output_column = "C"
        self.excel_power_voltage_column = "D"
        self.excel_power_current_column = "E"
        self.excel_power_meter_probe_column = "F"
        self.excel_aging_column = "G"
        self.excel_source_com_column = "H"
        self.excel_power_com_column = "I"
        self.excel_power_meter_com_column = "J"
        self.excel_spectrum_column = "K"
        self.excel_network_com_column = "L"
        self.excel_select_freq_column = "M"
        self.excel_output_off_ref_column = "N"
        self.excel_atr_start_input_column = "O"
        self.excel_p_sat_input_column = "P"
        self.excel_overdrive_input_column = "Q"
        self.excel_pout_column = "R"
        self.save_hor = True


class Excel(ExcelCommonVar, ExcelVariableDeclare):
    def __init__(self, path=""):
        ExcelCommonVar.__init__(self)
        ExcelVariableDeclare.__init__(self, path)

    def load_power_atr_excel_procedure(self, load_path):
        self.path = load_path
        self.__load_xls_sheet_all()
        self.__get_offset_procedure_version_0()

    def set_path(self, path):
        self.path = path

    def __load_xls_sheet_all(self):
        self.l_wb = load_workbook(self.path, data_only=True)
        ws_name = self.l_wb.get_sheet_names()

        while len(self.l_ws_list) != 0:
            self.l_ws_list.pop()
            print("clean l_ws_list")
        # Load all work sheet
        for i in range(0, len(ws_name)):
            self.l_ws_list.append(self.l_wb[ws_name[i]])

    # not finished
    # def load_xls_sheet_num(self, sheet_num):
    #     self.l_wb = load_workbook(self.path, data_only=True)
    #     ws_name = self.l_wb.get_sheet_names()
    #
    #     if (len(ws_name)-1 < sheet_num) or (sheet_num < 0):
    #         print("error in sheet num")
    #         self.load_clear = False
    #     else:
    #         self.l_ws_list[sheet_num] = self.l_wb[ws_name[sheet_num]]

    # frequency, io offset, voltage get from work sheet 0
    def __get_offset_procedure_version_0(self):
        # 현재 self.l_ws_list[0] frequency, input offset, output offset, voltage 정보가 들어 있다.
        # 이것을 추출하는 작업을 하려함.

        # 0. var reset
        # 1. freq config
        # 2. check input offset
        # 3. check output offset
        # 4. check power voltage set
        # 5. check power current limit
        # 6. check power meter probe channel
        # 7. check start aging time
        # 8. check source generate communication
        # 9. check power supply communication
        # 10. check power meter communication
        # 11. check spectrum communication
        # 12. check network communication
        # 13. check select frequency
        # 14. check output offset ref power
        # 15. check ATR start input dBm
        # 16. check P_sat input dBm
        # 17. check Overdrive input dBm
        # 18. check Pout output dBm
        # 19. load freq
        # 20. load input db
        # 21. load output db
        # 22. load Voltage set
        # 23. load Current limit
        # 24. load power meter probe channel
        # 25. load Start aging time
        # 26. load source generate communication
        # 27. load power supply communication
        # 28. load power meter communication
        # 29. load spectrum communication
        # 30. load network communication
        # 31. load select frequency
        # 32. load output offset ref power
        # 33. load ATR start input dBm
        # 34. load P_sat input dBm
        # 35. load Overdrive input dBm
        # 36. load Pout output dBm

        # 0. var reset
        self.__var_reset()
        # 1. freq config
        self.__get_unit_from_column(self.excel_freq_column)
        if self.load_clear is not True:
            return
        # 2. check input offset
        self.__get_unit_from_column(self.excel_input_column)
        if self.load_clear is not True:
            return
        # 3. check output offset
        self.__get_unit_from_column(self.excel_output_column)
        if self.load_clear is not True:
            return
        # 4. check power voltage set
        self.__get_unit_from_column(self.excel_power_voltage_column)
        if self.load_clear is not True:
            return
        # 5. check power current limit
        self.__get_unit_from_column(self.excel_power_current_column)
        if self.load_clear is not True:
            return
        # 6. check power meter probe channel
        self.__get_unit_from_column(self.excel_power_meter_probe_column)
        if self.load_clear is not True:
            return
        # 7. check start aging time
        self.__get_unit_from_column(self.excel_aging_column)
        if self.load_clear is not True:
            return
        # 8. check source generate communication
        self.__get_unit_from_column(self.excel_source_com_column)
        if self.load_clear is not True:
            return
        # 9. check power supply communication
        self.__get_unit_from_column(self.excel_power_com_column)
        if self.load_clear is not True:
            return
        # 10. check power meter communication
        self.__get_unit_from_column(self.excel_power_meter_com_column)
        if self.load_clear is not True:
            return
        # 11. check spectrum communication
        self.__get_unit_from_column(self.excel_spectrum_column)
        if self.load_clear is not True:
            return
        # 12. check network communication
        self.__get_unit_from_column(self.excel_network_com_column)
        if self.load_clear is not True:
            return
        # 13. check select frequency
        self.__get_unit_from_column(self.excel_select_freq_column)
        if self.load_clear is not True:
            return
        # 14. check output offset ref power
        self.__get_unit_from_column(self.excel_output_off_ref_column)
        if self.load_clear is not True:
            return
        # 15. check ATR start input dBm
        self.__get_unit_from_column(self.excel_atr_start_input_column)
        if self.load_clear is not True:
            return
        # 16. check P_sat input dBm
        self.__get_unit_from_column(self.excel_p_sat_input_column)
        if self.load_clear is not True:
            return
        # 17. check Overdrive input dBm
        self.__get_unit_from_column(self.excel_overdrive_input_column)
        if self.load_clear is not True:
            return
        # 18. check Pout output dBm
        self.__get_unit_from_column(self.excel_pout_column)
        if self.load_clear is not True:
            return
        # 19. load freq
        self.__load_var_from_column(self.excel_freq_column)
        # 20. load input db
        self.__load_var_from_column(self.excel_input_column)
        # 21. load output db
        self.__load_var_from_column(self.excel_output_column)
        # 22. load Voltage set
        self.__load_var_from_column(self.excel_power_voltage_column)
        # 23. load Current limit
        self.__load_var_from_column(self.excel_power_current_column)
        # 24. load power meter probe channel
        self.__load_var_from_column(self.excel_power_meter_probe_column)
        # 25. load Start aging time
        self.__load_var_from_column(self.excel_aging_column)
        # 26. load source generate communication
        self.__load_var_from_column(self.excel_source_com_column)
        # 27. load power supply communication
        self.__load_var_from_column(self.excel_power_com_column)
        # 28. load power meter communication
        self.__load_var_from_column(self.excel_power_meter_com_column)
        # 29. load spectrum communication
        self.__load_var_from_column(self.excel_spectrum_column)
        # 30. load network communication
        self.__load_var_from_column(self.excel_network_com_column)
        # 31. load select frequency
        self.__load_var_from_column(self.excel_select_freq_column)
        # 32. load output offset ref power
        self.__load_var_from_column(self.excel_output_off_ref_column)
        # 33. load ATR start input dBm
        self.__load_var_from_column(self.excel_atr_start_input_column)
        # 34. load P_sat input dBm
        self.__load_var_from_column(self.excel_p_sat_input_column)
        # 35. load Overdrive input dBm
        self.__load_var_from_column(self.excel_overdrive_input_column)
        # 36. load Pout output dBm
        self.__load_var_from_column(self.excel_pout_column)

    def __var_reset(self):
        self.freq_var = []
        self.input_offset_var = []
        self.output_offset_var = []
        self.power_voltage_var = ""
        self.power_current_var = ""
        self.power_cur_var = ""
        self.power_m_probe = ""
        self.aging_var = ""
        self.aging_left_var = ""
        self.aging_done_var = False
        self.reference_done = False
        self.source_com_var = ""
        self.power_com_var = ""
        self.power_meter_com_var = ""

        self.multiple_freq = 0  # Hz, KHz, MHz, GHz, THz
        self.multiple_aging = 0  # sec, min, hour
        self.source_comm_option = ""  # GPIB, USB, Serial
        self.power_comm_option = ""  # GPIB, USB, Serial
        self.power_meter_comm_option = ""  # GPIB, USB, Serial

        self.multiple_select_freq = 0  # Hz, KHz, MHz, GHz, THz
        self.select_freq_var = []  # select freq var
        self.output_offset_ref_power = ""  # select freq var
        self.atr_start_input_var = ""
        self.p_sat_input_var = ""
        self.overdrive_input_var = ""
        self.pout_var = ""

        self.freq_cell_name = ""  # default Frequency
        self.input_cell_name = ""  # default Input offset
        self.output_cell_name = ""  # default Output offset
        self.power_voltage_cell_name = ""  # default Power Voltage
        self.power_current_cell_name = ""  # default Power Voltage
        self.power_meter_probe_cell_name = ""  # default Power Voltage
        self.aging_cell_name = ""  # default Aging time
        self.source_comm_cell_name = ""  # default Source generate
        self.power_comm_cell_name = ""  # default Power supply
        self.power_meter_comm_cell_name = ""  # default Power meter
        self.select_freq_cell_name = ""  # default Select Freq
        self.atr_start_input_dbm_name = ""  # default ATR start input dBm
        self.p_sat_input_name = ""  # default P_sat input
        self.overdrive_input_name = ""  # default Overdrive input
        self.pout_name = ""  # default Pout
        self.load_clear = True

    def __get_unit_from_column(self, column):
        d_name_cell = column + str(self.excel_data_name_row)
        d_unit_cell = column + str(self.excel_data_unit_row)
        if column == self.excel_freq_column:
            self.multiple_freq = 0
            self.freq_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value == self.excel_freq_unit[0]:
                self.multiple_freq = 1
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_freq_unit[1]:
                self.multiple_freq = 1 * 1000
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_freq_unit[2]:
                self.multiple_freq = 1 * 1000 * 1000
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_freq_unit[3]:
                self.multiple_freq = 1 * 1000 * 1000 * 1000
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_freq_unit[4]:
                self.multiple_freq = 1 * 1000 * 1000 * 1000 * 1000
            else:
                pg.alert(text="freq unit 설정되지 않음\n" + "ex) Hz, KHz, MHz, GHz, THz",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_input_column:
            self.input_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value != self.excel_input_unit[0]:
                pg.alert(text="input offset unit 설정되지 않음\n" + "ex) dB",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_output_column:
            self.output_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value != self.excel_output_unit[0]:
                pg.alert(text="output offset unit 설정되지 않음\n" + "ex) dB",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_power_voltage_column:
            self.power_voltage_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value != self.excel_power_unit[0]:
                pg.alert(text="Power voltage unit 설정되지 않음\n" + "ex) V",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_power_current_column:
            self.power_current_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value != self.excel_current_unit[0]:
                pg.alert(text="Power current unit 설정되지 않음\n" + "ex) A",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_power_meter_probe_column:
            self.power_meter_probe_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value != self.excel_power_meter_probe_unit[0]:
                pg.alert(text="Power meter probe channel 설정되지 않음\n" + "ex) Channel",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_aging_column:
            self.multiple_aging = 0
            self.aging_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value == self.excel_aging_unit[0]:
                self.multiple_aging = 1
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_aging_unit[1]:
                self.multiple_aging = 1 * 60
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_aging_unit[2]:
                self.multiple_aging = 1 * 60 * 60
            else:
                pg.alert(text="Start aging time unit 설정되지 않음\n" + "ex) sec, min, hour",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_source_com_column:
            self.source_comm_option = ""
            self.source_comm_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[0]:
                self.source_comm_option = self.excel_com_unit[0]  # GPIB
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[1]:
                self.source_comm_option = self.excel_com_unit[1]  # USB
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[2]:
                self.source_comm_option = self.excel_com_unit[2]  # Serial
                self.source_baud_rate = self.l_ws_list[0][d_unit_cell + 1].value
            else:
                pg.alert(text="Source communication unit 설정되지 않음\n"
                              + "ex) GPIB, USB, Serial\nSerial baud rate 필요",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_power_com_column:
            self.power_comm_option = ""
            self.power_comm_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[0]:
                self.power_comm_option = self.excel_com_unit[0]  # GPIB
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[1]:
                self.power_comm_option = self.excel_com_unit[1]  # USB
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[2]:
                self.power_comm_option = self.excel_com_unit[2]  # Serial
                self.source_baud_rate = self.l_ws_list[0][d_unit_cell + 1].value
            else:
                pg.alert(text="Power supply communication unit 설정되지 않음\n"
                              + "ex) GPIB, USB, Serial\nSerial baud rate 필요",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_power_meter_com_column:
            self.power_meter_comm_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[0]:
                self.power_meter_comm_option = self.excel_com_unit[0]  # GPIB
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[1]:
                self.power_meter_comm_option = self.excel_com_unit[1]  # USB
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[2]:
                self.power_meter_comm_option = self.excel_com_unit[2]  # Serial
            else:
                pg.alert(text="Power meter communication unit 설정되지 않음\n"
                              + "ex) GPIB, USB, Serial\nSerial baud rate 필요",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_spectrum_column:
            self.spectrum_comm_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[0]:
                self.power_meter_comm_option = self.excel_com_unit[0]  # GPIB
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[1]:
                self.power_meter_comm_option = self.excel_com_unit[1]  # USB
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[2]:
                self.power_meter_comm_option = self.excel_com_unit[2]  # Serial
            else:
                pg.alert(text="Spectrum communication unit 설정되지 않음\n"
                              + "ex) GPIB, USB, Serial\nSerial baud rate 필요",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_network_com_column:
            self.network_comm_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[0]:
                self.power_meter_comm_option = self.excel_com_unit[0]  # GPIB
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[1]:
                self.power_meter_comm_option = self.excel_com_unit[1]  # USB
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_com_unit[2]:
                self.power_meter_comm_option = self.excel_com_unit[2]  # Serial
            else:
                pg.alert(text="Network communication unit 설정되지 않음\n"
                              + "ex) GPIB, USB, Serial\nSerial baud rate 필요",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_select_freq_column:
            self.multiple_select_freq = 0
            self.select_freq_cell_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value == self.excel_freq_unit[0]:
                self.multiple_select_freq = 1
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_freq_unit[1]:
                self.multiple_select_freq = 1 * 1000
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_freq_unit[2]:
                self.multiple_select_freq = 1 * 1000 * 1000
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_freq_unit[3]:
                self.multiple_select_freq = 1 * 1000 * 1000 * 1000
            elif self.l_ws_list[0][d_unit_cell].value == self.excel_freq_unit[4]:
                self.multiple_select_freq = 1 * 1000 * 1000 * 1000 * 1000
            else:
                pg.alert(text="Select freq unit 설정되지 않음\n" + "ex) Hz, KHz, MHz, GHz, THz",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_output_off_ref_column:
            self.output_off_ref_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value != self.excel_dbm_unit[0]:
                pg.alert(text="p_sat input dBm unit 설정되지 않음\n" + "ex) dBm",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_atr_start_input_column:
            self.atr_start_input_dbm_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value != self.excel_dbm_unit[0]:
                pg.alert(text="atr start dBm unit 설정되지 않음\n" + "ex) dBm",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_p_sat_input_column:
            self.output_off_ref_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value != self.excel_dbm_unit[0]:
                pg.alert(text="p_sat input dBm unit 설정되지 않음\n" + "ex) dBm",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_overdrive_input_column:
            self.overdrive_input_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value != self.excel_dbm_unit[0]:
                print("Probe channel not defined")
                pg.alert(text="overdrive dBm unit 설정되지 않음\n" + "ex) dBm",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        elif column == self.excel_pout_column:
            self.pout_name = d_name_cell
            if self.l_ws_list[0][d_unit_cell].value != self.excel_dbm_unit[0]:
                print("Probe channel not defined")
                pg.alert(text="overdrive dBm unit 설정되지 않음\n" + "ex) dBm",
                         title="Error",
                         button="확인")
                self.load_clear = False
                return
        else:
            print("알수없는 communication column")
            self.load_clear = False

    def __load_var_from_column(self, data_column):
        i = self.excel_data_start_row

        _data_cell = data_column + str(i)
        while self.l_ws_list[0][_data_cell].value is not None:
            print(self.l_ws_list[0][_data_cell].value)
            # check column
            if data_column == self.excel_freq_column:
                self.freq_var.append(self.l_ws_list[0][_data_cell].value * self.multiple_freq)
            elif data_column == self.excel_input_column:
                self.input_offset_var.append(self.l_ws_list[0][_data_cell].value)
            elif data_column == self.excel_output_column:
                self.output_offset_var.append(self.l_ws_list[0][_data_cell].value)
            elif data_column == self.excel_power_voltage_column:
                self.power_voltage_var = self.l_ws_list[0][_data_cell].value
            elif data_column == self.excel_power_current_column:
                self.power_cur_var = self.l_ws_list[0][_data_cell].value
            elif data_column == self.excel_power_meter_probe_column:
                self.power_m_probe = self.l_ws_list[0][_data_cell].value
            elif data_column == self.excel_aging_column:
                self.aging_var = self.l_ws_list[0][_data_cell].value * self.multiple_aging
            elif data_column == self.excel_source_com_column:
                self.source_com_var = self.l_ws_list[0][_data_cell].value
            elif data_column == self.excel_power_com_column:
                self.power_com_var = self.l_ws_list[0][_data_cell].value
            elif data_column == self.excel_power_meter_com_column:
                self.power_meter_com_var = self.l_ws_list[0][_data_cell].value
            elif data_column == self.excel_spectrum_column:
                self.spectrum_com_var = self.l_ws_list[0][_data_cell].value
            elif data_column == self.excel_network_com_column:
                self.network_com_var = self.l_ws_list[0][_data_cell].value
            elif data_column == self.excel_select_freq_column:
                self.select_freq_var.append(self.l_ws_list[0][_data_cell].value * self.multiple_select_freq)
            elif data_column == self.excel_output_off_ref_column:
                self.output_offset_ref_power = self.l_ws_list[0][_data_cell].value
            elif data_column == self.excel_atr_start_input_column:
                self.atr_start_input_var = self.l_ws_list[0][_data_cell].value
            elif data_column == self.excel_p_sat_input_column:
                self.p_sat_input_var = self.l_ws_list[0][_data_cell].value
            elif data_column == self.excel_overdrive_input_column:
                self.overdrive_input_var = self.l_ws_list[0][_data_cell].value
            elif data_column == self.excel_pout_column:
                self.pout_var = self.l_ws_list[0][_data_cell].value
            else:
                print("Error 등록되지 않은 id")
                self.load_clear = False
                return
            i += 1
            _data_cell = data_column + str(i)


class Dialog(CommonDialogVar):
    def __init__(self):
        CommonDialogVar.__init__(self)

    def main_dialog_open(self):
        self.g_root = tkinter.Tk()
        self.g_root.protocol("WM_DELETE_WINDOW", self.root_close)
        self.__make_main_panel()
        self.g_root.mainloop()

    def __make_main_panel(self):
        self.g_root.title("Main Dialog")
        self.g_root.geometry("640x350")
        self.g_root.resizable(False, False)
        self.g_root.iconbitmap("exodus.ico")  # exe 파일 폴더에 .ico 파일을 넣어주어야 한다.

        # remote power meter button
        self.__add_remote("Open", self.remote_p_meter_id)
        # remote power button
        self.__add_remote("Open", self.remote_power_id)
        # remote network button
        self.__add_remote("Open", self.remote_network_id)
        # remote source button
        self.__add_remote("Open", self.remote_source_id)
        # remote spectrum button
        self.__add_remote("Open", self.remote_spectrum_id)
        # remote all button
        self.__add_remote("Open", self.remote_all_id)

        # atr power button
        self.__add_atr("Open", self.atr_power_id)
        # atr network button
        self.__add_atr("Open", self.atr_network_id)
        # atr spectrum button
        self.__add_atr("Open", self.atr_spectrum_id)
        # atr offset button
        self.__add_atr("Open", self.atr_offset_id)

    def __add_remote(self, text, _id):
        label_column = 1
        button_column = 2

        button = tkinter.Button(self.g_root,
                                text=text,
                                command=partial(self.__button_clicked, _id),
                                width=self.main_butt_width,
                                height=self.main_butt_height)
        if _id is None:
            pass
        elif _id == self.remote_p_meter_id:
            _row = 1
            # label
            label = tkinter.Label(self.g_root, text="Power Meter Remote", anchor="w")
            label.grid(row=_row, column=label_column, sticky="NEWS",
                       padx=self.remote_start_x, pady=self.remote_start_y)
            # button
            button.grid(row=_row, column=button_column, sticky="NEWS",
                        padx=self.remote_button_x, pady=self.remote_start_y)
        elif _id == self.remote_power_id:
            _row = 2
            # label
            label = tkinter.Label(self.g_root, text="Power Remote", anchor="w")
            label.grid(row=_row, column=label_column, sticky="NEWS",
                       padx=self.remote_start_x, pady=self.remote_y_gap)
            # button
            button.grid(row=_row, column=button_column, sticky="NEWS",
                        padx=self.remote_button_x, pady=self.remote_y_gap)
        elif _id == self.remote_network_id:
            _row = 3
            # label
            label = tkinter.Label(self.g_root, text="Network Remote", anchor="w")
            label.grid(row=_row, column=label_column, sticky="NEWS",
                       padx=self.remote_start_x, pady=self.remote_y_gap)
            # button
            button.grid(row=_row, column=button_column, sticky="NEWS",
                        padx=self.remote_button_x, pady=self.remote_y_gap)
        elif _id == self.remote_source_id:
            _row = 4
            # label
            label = tkinter.Label(self.g_root, text="Source Remote", anchor="w")
            label.grid(row=_row, column=label_column, sticky="NEWS",
                       padx=self.remote_start_x, pady=self.remote_y_gap)
            # button
            button.grid(row=_row, column=button_column, sticky="NEWS",
                        padx=self.remote_button_x, pady=self.remote_y_gap)
        elif _id == self.remote_spectrum_id:
            _row = 5
            # label
            label = tkinter.Label(self.g_root, text="Spectrum Remote", anchor="w")
            label.grid(row=_row, column=label_column, sticky="NEWS",
                       padx=self.remote_start_x, pady=self.remote_y_gap)
            # button
            button.grid(row=_row, column=button_column, sticky="NEWS",
                        padx=self.remote_button_x, pady=self.remote_y_gap)
        elif _id == self.remote_all_id:
            _row = 6
            # label
            label = tkinter.Label(self.g_root, text="All Remote", anchor="w")
            label.grid(row=_row, column=label_column, sticky="NEWS",
                       padx=self.remote_start_x, pady=self.remote_y_gap)
            # button
            button.grid(row=_row, column=button_column, sticky="NEWS",
                        padx=self.remote_button_x, pady=self.remote_y_gap)
        else:
            pass

    def __add_atr(self, text, _id):
        label_column = 3
        button_column = 4

        button = tkinter.Button(self.g_root,
                                text=text,
                                command=partial(self.__button_clicked,
                                                _id),
                                width=self.main_butt_width,
                                height=self.main_butt_height)
        if _id == self.atr_power_id:
            _row = 1
            # label
            label = tkinter.Label(self.g_root, text="Power Meter ATR", anchor="w")
            label.grid(row=_row, column=label_column, sticky="NEWS",
                       padx=self.atr_start_x, pady=self.atr_start_y)
            # button
            button.grid(row=_row, column=button_column, sticky="NEWS",
                        padx=self.atr_button_x, pady=self.atr_start_y)
        elif _id == self.atr_network_id:
            _row = 2
            # label
            label = tkinter.Label(self.g_root, text="Network ATR", anchor="w")
            label.grid(row=_row, column=label_column, sticky="NEWS",
                       padx=self.atr_start_x, pady=self.atr_y_gap)
            # button
            button.grid(row=_row, column=button_column, sticky="NEWS",
                        padx=self.atr_button_x, pady=self.atr_y_gap)
        elif _id == self.atr_spectrum_id:
            _row = 3
            # label
            label = tkinter.Label(self.g_root, text="Spectrum ATR", anchor="w")
            label.grid(row=_row, column=label_column, sticky="NEWS",
                       padx=self.atr_start_x, pady=self.atr_y_gap)
            # button
            button.grid(row=_row, column=button_column, sticky="NEWS",
                        padx=self.atr_button_x, pady=self.atr_y_gap)
        elif _id == self.atr_offset_id:
            _row = 4
            # label
            label = tkinter.Label(self.g_root, text="Offset", anchor="w")
            label.grid(row=_row, column=label_column, sticky="NEWS",
                       padx=self.atr_start_x, pady=self.atr_y_gap)
            # button
            button.grid(row=_row, column=button_column, sticky="NEWS",
                        padx=self.atr_button_x, pady=self.atr_y_gap)
        else:
            pass

    def __button_clicked(self, _id):
        if _id is None:
            pass
        elif _id == self.remote_p_meter_id:
            pass
        elif _id == self.remote_power_id:
            pass
        elif _id == self.remote_network_id:
            pass
        elif _id == self.remote_source_id:
            pass
        elif _id == self.remote_spectrum_id:
            pass
        elif _id == self.remote_all_id:
            pass
        elif _id == self.atr_power_id:
            # load excel dialog open
            self.load_file_dialog()
            if self.excel_path != "-1":
                self.excel.load_power_atr_excel_procedure(load_path=self.excel_path)
                if self.excel.load_clear is True:
                    self.__new_window(_id=self.atr_power_id)
        elif _id == self.atr_network_id:
            pass
        elif _id == self.atr_spectrum_id:
            pass
        elif _id == self.atr_offset_id:
            self.load_file_dialog()
            if self.excel_path != "-1":
                self.excel.load_power_atr_excel_procedure(load_path=self.excel_path)
                if self.excel.load_clear is True:
                    self.__new_window(_id=self.atr_offset_id)

    def __atr_power_new_window(self):
        new_window = tkinter.Toplevel(self.g_root)
        new_window.geometry("800x500")
        new_window.title("Power ATR")
        new_window.iconbitmap("exodus.ico")
        new_window.resizable(False, False)

        entry_width = 15

        # line_0
        # panel label
        label_panel = tkinter.Label(new_window, text="", anchor="w")
        label_panel.grid(row=0, column=0, columnspan=5, sticky="NEWS", padx=10, pady=10)
        # line_1
        # frequency label
        label = tkinter.Label(new_window, text="Frequency")
        label.grid(row=1, column=0, sticky="NEWS")
        # input offset label
        label = tkinter.Label(new_window, text="Input offset")
        label.grid(row=1, column=1, sticky="NEWS")
        # output offset label
        label = tkinter.Label(new_window, text="Output offset")
        label.grid(row=1, column=2, sticky="NEWS")
        # Power voltage label
        label = tkinter.Label(new_window, text="Power voltage")
        label.grid(row=1, column=3, sticky="NEWS")
        # Power current label
        label = tkinter.Label(new_window, text="Power current")
        label.grid(row=1, column=4, sticky="NEWS")
        # Aging time label
        label = tkinter.Label(new_window, text="Aging time")
        label.grid(row=1, column=5, sticky="NEWS")

        # line_2
        # frequency entry
        self.atr_freq = tkinter.StringVar(new_window)
        if self.excel.freq_var[0] >= 1000000000000:
            self.atr_freq.set("{0} THz".format(self.excel.freq_var[0] / 1000000000000))
        elif self.excel.freq_var[0] >= 1000000000:
            self.atr_freq.set("{0} GHz".format(self.excel.freq_var[0] / 1000000000))
        elif self.excel.freq_var[0] >= 1000000:
            self.atr_freq.set("{0} MHz".format(self.excel.freq_var[0] / 1000000))
        elif self.excel.freq_var[0] >= 1000:
            self.atr_freq.set("{0} KHz".format(self.excel.freq_var[0] / 1000))
        else:
            self.atr_freq.set("{0} Hz".format(self.excel.freq_var[0]))
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_freq,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=2, column=0, sticky="NEWS")
        # input offset entry
        self.atr_input_offset = tkinter.StringVar(new_window)
        self.atr_input_offset.set("{0} dB".format(self.excel.input_offset_var[0]))
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_input_offset,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=2, column=1, sticky="NEWS")
        # output offset entry
        self.atr_output_offset = tkinter.StringVar(new_window)
        self.atr_output_offset.set("{0} dB".format(self.excel.output_offset_var[0]))
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_output_offset,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=2, column=2, sticky="NEWS")
        # Power voltage entry
        self.atr_power_voltage = tkinter.StringVar(new_window)
        self.atr_power_voltage.set("{0} V".format(self.excel.power_voltage_var))
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_power_voltage,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=2, column=3, sticky="NEWS")
        # Power current entry
        self.atr_power_current = tkinter.StringVar(new_window)
        self.atr_power_current.set("{0} A".format(self.excel.power_current_var))
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_power_current,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=2, column=4, sticky="NEWS")
        # Aging time entry
        self.atr_aging_time = tkinter.StringVar(new_window)
        self.atr_aging_time.set("{0} sec".format(self.excel.aging_var))
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_aging_time,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=2, column=5, sticky="NEWS")

        # line_3
        # rf input label
        label = tkinter.Label(new_window, text="RF Input     ", anchor="e")
        label.grid(row=3, column=0, sticky="NEWS")
        # rf output label
        label = tkinter.Label(new_window, text="RF Output    ", anchor="e")
        label.grid(row=3, column=1, sticky="NEWS")
        # current label
        label = tkinter.Label(new_window, text="Current")
        label.grid(row=3, column=2, sticky="NEWS")
        # aging time left label
        label = tkinter.Label(new_window, text="Aging left")
        label.grid(row=3, column=3, sticky="NEWS")

        # line_4
        # RF input entry
        self.atr_rf_input = tkinter.StringVar(new_window)
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_rf_input,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=4, column=0, sticky="NES")
        # RF output entry
        self.atr_rf_output = tkinter.StringVar(new_window)
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_rf_output,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=4, column=1, sticky="NES")
        # Current entry
        self.atr_current = tkinter.StringVar(new_window)
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_current,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=4, column=2, sticky="NEWS")
        # aging time left entry
        self.atr_aging_left = tkinter.StringVar(new_window)
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_aging_left,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=4, column=3, sticky="NEWS")

        # line_5
        # rf output label
        label = tkinter.Label(new_window, text="  System option", anchor="w")
        label.grid(row=5, column=0, sticky="NEWS", columnspan=2)
        # System Forward Label
        label = tkinter.Label(new_window, text="Sys Fwd")
        label.grid(row=5, column=2, sticky="NEWS")
        # System Frequency Label
        label = tkinter.Label(new_window, text="Sys Freq")
        label.grid(row=5, column=3, sticky="NEWS")
        # System Temperature Label
        label = tkinter.Label(new_window, text="Sys Temp")
        label.grid(row=5, column=4, sticky="NEWS")

        # line_6
        # sys option checkbox
        checkbox = tkinter.Checkbutton(new_window,
                                       text="                      ",
                                       command=self.atr_sys_opt_clicked)
        checkbox.grid(row=6, column=0, columnspan=2, sticky="NEWS")
        # sys fwd entry
        self.atr_sys_fwd = tkinter.StringVar(new_window)
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_sys_fwd,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=6, column=2, sticky="EW")
        # sys freq entry
        self.atr_sys_freq = tkinter.StringVar(new_window)
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_sys_freq,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=6, column=3, sticky="EW")
        # sys freq temp
        self.atr_sys_temp = tkinter.StringVar(new_window)
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_sys_temp,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=6, column=4, sticky="EW")

        # line 7
        # System cmd Label
        label = tkinter.Label(new_window, text="Sys cmd", anchor="w")
        label.grid(row=7, column=0, sticky="NEWS", columnspan=2)
        # Source communication label
        label = tkinter.Label(new_window, text="Source com")
        label.grid(row=7, column=2, sticky="NEWS")
        # Power supply communication label
        label = tkinter.Label(new_window, text="Power com")
        label.grid(row=7, column=3, sticky="NEWS")
        # Power meter communication label
        label = tkinter.Label(new_window, text="Power_m com")
        label.grid(row=7, column=4, sticky="NEWS")

        # line 8
        # System cmd text
        self.atr_sys_cmd = tkinter.StringVar(new_window)
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_sys_cmd,
                              width=entry_width * 2,
                              justify="center")
        entry.grid(row=8, column=0, sticky="EW", columnspan=2)
        # source communication entry
        self.atr_source_com = tkinter.StringVar(new_window)
        self.atr_source_com.set(self.excel.source_com_var)
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_source_com,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=8, column=2, sticky="EW")
        # power supply communication entry
        self.atr_power_com = tkinter.StringVar(new_window)
        self.atr_power_com.set(self.excel.power_com_var)
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_power_com,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=8, column=3, sticky="EW")
        # power meter communication entry
        self.atr_power_meter_com = tkinter.StringVar(new_window)
        self.atr_power_meter_com.set(self.excel.power_meter_com_var)
        entry = tkinter.Entry(new_window,
                              textvariable=self.atr_power_meter_com,
                              width=entry_width,
                              state="readonly",
                              justify="center")
        entry.grid(row=8, column=4, sticky="EW")

        # line 9
        # ATR start button
        button = tkinter.Button(new_window,
                                text="START",
                                width=self.main_butt_width,
                                height=self.main_butt_height,
                                command=partial(self.__atr_button_clicked, self.atr_start_button_id))
        button.grid(row=9, column=0, sticky="NEWS", columnspan=2)
        # ATR stop button
        button = tkinter.Button(new_window,
                                text="STOP",
                                width=self.main_butt_width,
                                height=self.main_butt_height,
                                command=partial(self.__atr_button_clicked, self.atr_stop_button_id))
        button.grid(row=9, column=2, sticky="NEWS", columnspan=2)

        # line 10
        # ATR start button
        button = tkinter.Button(new_window,
                                text="DISPLAY",
                                width=self.main_butt_width,
                                height=self.main_butt_height,
                                command=partial(self.__atr_button_clicked, self.atr_display_button_id))
        button.grid(row=10, column=0, sticky="NEWS", columnspan=2)
        # ATR stop button
        button = tkinter.Button(new_window,
                                text="SAVE",
                                width=self.main_butt_width,
                                height=self.main_butt_height,
                                command=partial(self.__atr_button_clicked, self.atr_save_button_id))
        button.grid(row=10, column=2, sticky="NEWS", columnspan=2)

    def __offset_new_window(self):
        new_window = tkinter.Toplevel(self.g_root)
        new_window.geometry("350x200")
        new_window.title("Offset")
        new_window.iconbitmap("exodus.ico")
        new_window.resizable(False, False)

        # input offset start button
        button = tkinter.Button(new_window,
                                text="Measure Input",
                                width=self.main_butt_width * 3,
                                height=self.main_butt_height,
                                command=partial(self.__atr_button_clicked, self.atr_measure_input_offset_id))
        button.grid(row=0, column=0, sticky="NEWS", columnspan=1, padx=30, pady=15)

        # output offset start button
        button = tkinter.Button(new_window,
                                text="Measure Output",
                                width=self.main_butt_width * 3,
                                height=self.main_butt_height,
                                command=partial(self.__atr_button_clicked, self.atr_measure_output_offset_id))
        button.grid(row=1, column=0, sticky="NEWS", columnspan=1, padx=30, pady=15)

        # save all offset var button
        button = tkinter.Button(new_window,
                                text="Save All",
                                width=self.main_butt_width * 3,
                                height=self.main_butt_height,
                                command=partial(self.__atr_button_clicked, self.atr_save_all_offset_id))
        button.grid(row=2, column=0, sticky="NEWS", columnspan=1, padx=30, pady=15)

    def __new_window(self, _id):
        if _id == self.remote_power_id:
            pass
        elif _id == self.remote_network_id:
            pass
        elif _id == self.remote_source_id:
            pass
        elif _id == self.remote_spectrum_id:
            pass
        elif _id == self.remote_all_id:
            pass
        elif _id == self.atr_power_id:
            self.__atr_power_new_window()
        elif _id == self.atr_network_id:
            pass
        elif _id == self.atr_spectrum_id:
            pass
        elif _id == self.atr_offset_id:
            self.__offset_new_window()

    def __atr_button_clicked(self, _id):
        if _id == self.atr_start_button_id:
            self.atr_start_clicked()
        elif _id == self.atr_stop_button_id:
            self.atr_stop_clicked()
        elif _id == self.atr_display_button_id:
            self.atr_display_clicked()
        elif _id == self.atr_save_button_id:
            # save excel dialog open
            self.save_file_dialog(_id=self.atr_save_button_id)
            if self.excel_path != "-1":
                self.save_all_var(save_path=self.excel_path)
        elif _id == self.atr_measure_input_offset_id:
            self.measure_offset_clicked(_id=_id)
        elif _id == self.atr_measure_output_offset_id:
            self.measure_offset_clicked(_id=_id)
        elif _id == self.atr_save_all_offset_id:
            self.save_all_offset_clicked()
        else:
            print("atr button id error")

    def offset_procedure_input(self):
        source_power = 0
        under_zero = 1
        # 0. set var
        # 1. set instrument instance
        # 2. reset var
        # 3. reset instrument
        # 4. set instrument

        # 0. set var
        _send_freq = int(self.excel.select_freq_var[self.atr_index])
        _send_offset_i = 0
        _send_offset_o = 0
        if _send_offset_i is not None:
            if _send_offset_i < 0:
                _send_offset_i = -_send_offset_i
        if _send_offset_o is not None:
            if _send_offset_o < 0:
                _send_offset_o = -_send_offset_o
        # 1. set instrument
        self.atr_seq += 1
        if self.atr_seq == 1:
            # 1. set instrument instance
            self.__atr_inst_call()
            # 2. reset var
            self._reset_input_offset_var()
        # 3. reset instrument
        elif self.atr_seq == 2:
            self.inst_source.set_output_agilent(on_off=False)  # set source off
            self.inst_power_meter.set_rel_agilent(on_off=False)  # set power meter rel off
        # 4. set instrument
        elif self.atr_seq == 3:
            self.inst_source.set_offset_agilent(offset=0)  # set source offset
            self.inst_power_meter.set_offset_agilent(offset=0, ch=self.excel.power_m_probe)  # set power meter offset
        elif self.atr_seq == 4:
            self.inst_source.set_dbm_agilent(dbm=source_power)  # set dBm
        elif self.atr_seq == 5:
            self.inst_source.set_output_agilent(on_off=True)  # set source on
        elif self.atr_seq == 6:
            self.inst_source.set_freq_agilent(freq=_send_freq)  # set source frequency
            self.inst_power_meter.set_freq_agilent(freq=_send_freq, ch=self.excel.power_m_probe)  # set power_m freq
        elif self.atr_seq == 7:
            self.atr_compare_0 = self.inst_power_meter.get_output(display_ch=1, round_num=2)
        elif self.atr_seq == 8:
            self.atr_compare_1 = self.inst_power_meter.get_output(display_ch=1, round_num=2)
        elif self.atr_seq == 9:
            if abs(self.atr_compare_0 - self.atr_compare_1) > 1e-9:
                self.atr_seq -= 3
            else:
                self.mes_input_offset.append(self.inst_power_meter.get_output(display_ch=1, round_num=under_zero))
                if len(self.mes_input_offset) == len(self.excel.select_freq_var):
                    # save excel dialog open
                    self.inst_source.set_output_agilent(on_off=False)  # set source off
                    pg.alert(text="Done",
                             title="Done",
                             button="확인")
                    return
                else:
                    self.atr_index += 1
                    self.atr_seq -= 4
        self.after_call(ms=250, func=self.offset_procedure_input)

    def offset_procedure_output(self):
        source_power = self.excel.output_offset_ref_power
        under_zero = 2
        # 0. set var
        # 1. set instrument instance
        # 2. reset var
        # 3. reset instrument
        # 4. set instrument

        # 0. set var
        _send_freq = int(self.excel.select_freq_var[self.atr_index])
        _send_offset_i = 0
        _send_offset_o = 0
        if _send_offset_i is not None:
            if _send_offset_i < 0:
                _send_offset_i = -_send_offset_i
        if _send_offset_o is not None:
            if _send_offset_o < 0:
                _send_offset_o = -_send_offset_o
        # 1. set instrument
        self.atr_seq += 1
        if self.atr_seq == 1:
            # 1. set instrument instance
            self.__atr_inst_call()
            # 2. reset var
            self._reset_output_offset_var()
        # 3. reset instrument
        elif self.atr_seq == 2:
            self.inst_source.set_output_agilent(on_off=False)  # set source off
            self.inst_power_meter.set_rel_agilent(on_off=False)  # set power meter rel off
        # 4. set instrument
        elif self.atr_seq == 3:
            self.inst_source.set_offset_agilent(offset=0)  # set source offset
            self.inst_power_meter.set_offset_agilent(offset=0, ch=self.excel.power_m_probe)  # set power meter offset
        elif self.atr_seq == 4:
            self.inst_source.set_dbm_agilent(dbm=source_power)  # set dBm
        elif self.atr_seq == 5:
            self.inst_source.set_output_agilent(on_off=True)  # set source on
        elif self.atr_seq == 6:
            self.inst_source.set_freq_agilent(freq=_send_freq)  # set source frequency
            self.inst_power_meter.set_freq_agilent(freq=_send_freq, ch=self.excel.power_m_probe)  # set power_m freq
        elif self.atr_seq == 7:
            self.atr_compare_0 = self.inst_power_meter.get_output(display_ch=1, round_num=2)
        elif self.atr_seq == 8:
            self.atr_compare_1 = self.inst_power_meter.get_output(display_ch=1, round_num=2)
        elif self.atr_seq == 9:
            if abs(self.atr_compare_0 - self.atr_compare_1) > 1e-9:
                self.atr_seq -= 3
            else:
                self.mes_output_offset.append(round(self.inst_power_meter.get_output(display_ch=1, round_num=under_zero)
                                              + self.excel.output_offset_ref_power, 2))
                if len(self.mes_output_offset) == len(self.excel.select_freq_var):
                    self.inst_source.set_output_agilent(on_off=False)  # set source off
                    pg.alert(text="Done",
                             title="Done",
                             button="확인")
                    return
                else:
                    self.atr_index += 1
                    self.atr_seq -= 4
        self.after_call(ms=250, func=self.offset_procedure_output)

    # input offset False -> output offset
    def offset_start(self, input_offset=True):
        if input_offset:
            self.offset_procedure_input()
        else:
            self.offset_procedure_output()

    def save_all_offset_clicked(self):
        # save excel dialog open
        self.save_file_dialog(_id=self.atr_measure_output_offset_id)
        if self.excel_path != "-1":
            self.save_option(save_path=self.excel_path, _id=self.atr_save_all_offset_id)

    def measure_offset_clicked(self, _id):
        self.atr_seq = 0
        if _id == self.atr_measure_input_offset_id:
            self.offset_start(input_offset=True)
        elif _id == self.atr_measure_output_offset_id:
            self.offset_start(input_offset=False)

    def atr_start_clicked(self):
        self.atr_seq = 0
        self.atr_ready()  # set config, power on, rf on

    def atr_ready(self):
        # 0. set var
        # 1. set instrument instance
        # 2. reset atr var
        # 3. reset instrument
        # 4. set instrument

        # 0. set var
        _send_freq = int(self.excel.select_freq_var[self.atr_index])
        _send_offset_i = float(self.__find_offset_in_table(send_freq=_send_freq, select="input_offset"))
        _send_offset_o = float(self.__find_offset_in_table(send_freq=_send_freq, select="output_offset"))
        if _send_offset_i is not None:
            if _send_offset_i < 0:
                _send_offset_i = -_send_offset_i
        if _send_offset_o is not None:
            if _send_offset_o < 0:
                _send_offset_o = -_send_offset_o
        # set instrument
        self.atr_seq += 1
        if self.atr_seq == 1:
            # 1. set instrument instance
            self.__atr_inst_call()
            # 2. reset atr var
            self._reset_atr_var()
        elif self.atr_seq == 2:  # set config_0
            self.inst_source.set_output_agilent(on_off=False)  # turn off source
            self.inst_power_meter.set_rel_agilent(on_off=False)  # reference off
            # test for power except
            self.inst_power.set_output_hp_6x74a(on_off=False)  # turn off power

        elif self.atr_seq == 3:  # set config_1
            self.inst_power.set_voltage_hp_6x74a(voltage=self.excel.power_voltage_var)  # power voltage set
            self.dialog_var_set(var_name="atr_power_voltage", value=self.excel.power_voltage_var)  # set dialog var
            self.inst_source.set_freq_agilent(freq=_send_freq)  # set frequency
            self.inst_power_meter.set_freq_agilent(freq=_send_freq)  # set frequency
            self.dialog_var_set(var_name="atr_freq", value=_send_freq)  # set dialog var
        elif self.atr_seq == 4:  # set config_2
            self.inst_power.set_current_hp_6x74a(current=self.excel.power_cur_var)  # power current set
            self.dialog_var_set(var_name="atr_power_current", value=self.excel.power_cur_var)  # set dialog var
            self.inst_source.set_offset_agilent(offset=_send_offset_i)  # set offset
            self.inst_power_meter.set_offset_agilent(offset=_send_offset_o, ch=self.excel.power_m_probe)  # set loss
            self.dialog_var_set(var_name="atr_input_offset", value=_send_offset_i)  # set dialog var
            self.dialog_var_set(var_name="atr_output_offset", value=_send_offset_o)  # set dialog var
        elif self.atr_seq == 5:  # set config_3
            self.inst_source.set_dbm_agilent(self.excel.atr_start_input_var)  # set dBm
            self.dialog_var_set(var_name="atr_rf_input", value=self.excel.atr_start_input_var)  # set dialog var
        elif self.atr_seq == 6:  # release_0
            # test for power except
            self.inst_power.set_output_hp_6x74a(on_off=True)  # turn on power

        elif self.atr_seq == 7:  # release_1
            self.inst_source.set_output_agilent(on_off=True)  # turn on rf state
        else:
            self.atr_seq = 0
            self.after_call(ms=250, func=self.aging)
            return
        self.after_call(ms=250, func=self.atr_ready)

    def aging(self):
        if int(self.excel.aging_left_var) > 0:
            self.atr_seq += 1
            if self.atr_seq == 1:
                self.inst_source.set_dbm_agilent(dbm=self.excel.atr_start_input_var)  # set dBm
                self.atr_input_input = self.excel.atr_start_input_var
                self.dialog_var_set(var_name="atr_rf_input", value=self.excel.atr_start_input_var)
                self.after_call(ms=250, func=self.aging)
            elif self.atr_seq == 2:
                self.atr_compare_0 = self.inst_power_meter.get_output(display_ch=1, round_num=2)
                self.dialog_var_set(var_name="atr_rf_output", value=self.atr_compare_0)
                self.after_call(ms=250, func=self.aging)
            elif self.atr_seq == 3:
                self.atr_compare_1 = self.inst_power_meter.get_output(display_ch=1, round_num=2)
                self.dialog_var_set(var_name="atr_rf_output", value=self.atr_compare_1)
                self.after_call(ms=250, func=self.aging)
            # 4. input get
            elif self.atr_seq == 4:
                if abs(self.atr_compare_0 - self.atr_compare_1) > 1e-9:
                    self.atr_seq -= 3
                else:
                    self.adder = MyCal.get_input_adder(
                        _output_now=self.atr_compare_0, _output_goal=self.excel.pout_var)
                    if self.adder is not None:
                        if self.adder != 0:
                            self.excel.aging_left_var = self.excel.aging_var
                            self.atr_input_input += self.adder
                            if self.atr_input_input <= self.excel.p_sat_input_var:
                                self.inst_source.set_dbm_agilent(dbm=self.atr_input_input)  # new input set
                                self.atr_seq -= 3
                                self.dialog_var_set(var_name="atr_rf_input", value=self.atr_input_input)
                                self.after_call(ms=250, func=self.aging)
                            else:
                                print("p input set over range")
                                return
                        else:
                            self.atr_seq -= 3
                            # set dialog var
                            self.excel.aging_left_var -= 1  # 1sec decrease
                            self.dialog_var_set(var_name="aging_time_left", value=self.excel.aging_left_var)
                            self.after_call(ms=500, func=self.aging)
                    else:
                        print("adder error")
                        return
        else:
            print("aging left is 0 or less")
            self.atr_seq = 0
            self.after_call(ms=250, func=self.atr_start)

    def atr_start(self):
        # 0 atr index check
        # 1. set var
        # 2. set instrument default
        # 3. p1 ready
        # 4. p1 get
        # 5. input ready
        # 6. input get
        # 7. p_sat ready
        # 8. p_sat get
        # 9. ready overdrive
        # 10. get overdrive
        # 11. atr index increase
        # 12. atr_seq increase
        # 13. groot after set

        # 1. set var
        _send_freq = int(self.excel.select_freq_var[self.atr_index])
        _send_offset_in = round(float(self.__find_offset_in_table(send_freq=_send_freq, select="input_offset")), 2)
        _send_offset_out = round(float(self.__find_offset_in_table(send_freq=_send_freq, select="output_offset")), 2)
        _send_input = self.excel.atr_start_input_var
        if (_send_offset_in is not None) and (_send_offset_out is not None):
            if _send_offset_in < 0:
                _send_offset_in = -_send_offset_in
            if _send_offset_out < 0:
                _send_offset_out = -_send_offset_out
        else:
            return
        # 2. set instrument default
        if self.atr_seq == 0:
            self.inst_source.set_output_agilent(on_off=False)  # turn off rf state
        # 3. p1 ready
        elif self.atr_seq == 1:
            self.inst_source.set_freq_agilent(freq=_send_freq)  # set frequency
            self.inst_power_meter.set_freq_agilent(ch=self.excel.power_m_probe, freq=_send_freq)
            self.dialog_var_set(var_name="atr_freq", value=_send_freq)  # set dialog var
        elif self.atr_seq == 2:
            self.inst_source.set_offset_agilent(offset=_send_offset_in)  # set offset
            self.inst_power_meter.set_offset_agilent(offset=_send_offset_out, ch=self.excel.power_m_probe)  # set loss
            self.dialog_var_set(var_name="atr_input_offset", value=_send_offset_in)  # set dialog var
            self.dialog_var_set(var_name="atr_output_offset", value=_send_offset_out)  # set dialog var
        elif self.atr_seq == 3:
            self.inst_source.set_dbm_agilent(dbm=self.excel.atr_start_input_var)  # set dBm
            self.dialog_var_set(var_name="atr_rf_input", value=self.excel.atr_start_input_var)
            self.inst_power_meter.set_rel_agilent(on_off=False)  # reference off
        elif self.atr_seq == 4:
            self.inst_source.set_output_agilent(on_off=True)  # turn on rf state
        elif self.atr_seq == 5:
            self.inst_power_meter.set_rel_agilent(on_off=True)  # reference on
            self.atr_p1_ref_input = 0
        # 2. p1 get
        elif self.atr_seq == 6:
            self.atr_compare_0 = self.inst_power_meter.get_rel(display_ch=1, round_num=2)
            self.dialog_var_set(var_name="atr_rf_output", value=self.atr_compare_0)
        elif self.atr_seq == 7:
            self.atr_compare_1 = self.inst_power_meter.get_rel(display_ch=1, round_num=2)
            self.dialog_var_set(var_name="atr_rf_output", value=self.atr_compare_0)
        elif self.atr_seq == 8:
            if abs(self.atr_compare_0 - self.atr_compare_1) > 1e-9:
                self.atr_seq -= 3
            else:
                self.adder = MyCal.get_p1_adder(
                    _out_ref=self.atr_compare_0, _input_ref=self.atr_p1_ref_input)
                send_var = self.excel.atr_start_input_var + self.atr_p1_ref_input
                if self.adder is not None:
                    if self.adder != 0:
                        self.atr_p1_ref_input += self.adder
                        send_var = self.excel.atr_start_input_var + self.atr_p1_ref_input
                        if send_var <= self.excel.p_sat_input_var:
                            self.inst_source.set_dbm_agilent(dbm=send_var)  # new input set
                            self.dialog_var_set(var_name="atr_rf_input", value=send_var)
                            self.atr_seq -= 3
                        else:
                            print("p1 input set over range")
                            self.inst_source.set_dbm_agilent(dbm=self.excel.atr_start_input_var)  # default input set
                            self.dialog_var_set(var_name="atr_rf_input", value=self.excel.atr_start_input_var)
                            return
                    else:
                        self.atr_p1_var.append(send_var)
                        self.inst_power_meter.set_rel_agilent(on_off=False)  # reference off
        # 3. input ready
        elif self.atr_seq == 9:
            self.inst_source.set_dbm_agilent(dbm=self.excel.atr_start_input_var)  # set dBm
            self.atr_input_input = self.excel.atr_start_input_var
            self.dialog_var_set(var_name="atr_rf_input", value=self.excel.atr_start_input_var)
        elif self.atr_seq == 10:
            self.atr_compare_0 = self.inst_power_meter.get_output(display_ch=1, round_num=2)
            self.dialog_var_set(var_name="atr_rf_output", value=self.atr_compare_0)
        elif self.atr_seq == 11:
            self.atr_compare_1 = self.inst_power_meter.get_output(display_ch=1, round_num=2)
            self.dialog_var_set(var_name="atr_rf_output", value=self.atr_compare_1)
        # 4. input get
        elif self.atr_seq == 12:
            if abs(self.atr_compare_0 - self.atr_compare_1) > 1e-9:
                self.atr_seq -= 3
            else:
                self.adder = MyCal.get_input_adder(_output_now=self.atr_compare_0, _output_goal=self.excel.pout_var)
                if self.adder is not None:
                    if self.adder != 0:
                        self.atr_input_input += self.adder
                        if self.atr_input_input <= self.excel.p_sat_input_var:
                            self.inst_source.set_dbm_agilent(self.atr_input_input)  # new input set
                            self.dialog_var_set(var_name="atr_rf_input", value=self.atr_input_input)
                            self.atr_seq -= 3
                        else:
                            print("p input set over range")
                            return
                    else:
                        self.atr_input_var.append(self.atr_input_input)
                        _current = self.inst_power.get_current_hp_6x74a(round_num=2)
                        self.atr_input_curr_var.append(_current)
                        self.dialog_var_set(var_name="atr_power_current", value=_current)
                else:
                    print("input adder error")
                    return
        # 5. p_sat ready
        elif self.atr_seq == 13:
            self.inst_source.set_dbm_agilent(dbm=self.excel.p_sat_input_var)  # set p_sat input dBm
            self.dialog_var_set(var_name="atr_rf_input", value=self.excel.p_sat_input_var)
        # 6. p_sat get
        elif self.atr_seq == 14:
            self.atr_compare_0 = self.inst_power_meter.get_output(display_ch=1, round_num=2)
            self.dialog_var_set(var_name="atr_rf_output", value=self.atr_compare_0)
        elif self.atr_seq == 15:
            self.atr_compare_1 = self.inst_power_meter.get_output(display_ch=1, round_num=2)
            self.dialog_var_set(var_name="atr_rf_output", value=self.atr_compare_1)
        elif self.atr_seq == 16:
            if abs(self.atr_compare_0 - self.atr_compare_1) > 1e-9:
                self.atr_seq -= 3
            else:
                self.atr_p_sat_var.append(self.atr_compare_0)
                # test for power except
                # load curr from instrument
                _current = self.inst_power.get_current_hp_6x74a(round_num=2)
                self.atr_p_sat_curr_var.append(_current)
                self.dialog_var_set(var_name="atr_power_current", value=_current)
        # 7. ready overdrive
        elif self.atr_seq == 17:
            self.inst_source.set_dbm_agilent(dbm=self.excel.overdrive_input_var)
            self.dialog_var_set(var_name="atr_rf_input", value=self.excel.overdrive_input_var)
        # 8. get overdrive
        elif self.atr_seq == 18:
            self.atr_compare_0 = self.inst_power_meter.get_output(display_ch=1, round_num=2)
            self.dialog_var_set(var_name="atr_rf_output", value=self.atr_compare_0)
        elif self.atr_seq == 19:
            self.atr_compare_1 = self.inst_power_meter.get_output(display_ch=1, round_num=2)
            self.dialog_var_set(var_name="atr_rf_output", value=self.atr_compare_1)
        elif self.atr_seq == 20:
            if abs(self.atr_compare_0 - self.atr_compare_1) > 1e-9:
                self.atr_seq -= 3
            else:
                self.atr_overdrive_var.append(self.atr_compare_0)
                # test for power except
                _current = self.inst_power.get_current_hp_6x74a(round_num=2)
                self.atr_overdrive_curr_var.append(_current)
                self.dialog_var_set(var_name="atr_power_current", value=_current)
        elif self.atr_seq == 21:
            if self.atr_index >= len(self.excel.select_freq_var) - 1:
                print("atr start end")
                self.atr_seq = 0
                self.after_call(ms=500, func=self.atr_end)
                return
            else:
                self.atr_seq = -1  # go to atr_seq == 0
                self.atr_index += 1
        else:
            print("unknown sequence number")
            return
        # 10. atr_seq increase
        self.atr_seq += 1
        # 11. groot after set
        self.after_call(ms=250, func=self.atr_start)

    def atr_end(self):
        # 1. reset instrument
        if self.atr_seq == 0:
            # test for power except
            self.inst_power.set_output_hp_6x74a(on_off=False)  # turn off power
        elif self.atr_seq == 1:
            self.inst_source.set_output_agilent(on_off=False)  # turn off rf state
        elif self.atr_seq == 2:
            self.inst_power_meter.set_rel_agilent(on_off=False)  # reference off
        else:
            print("atr end end")
            return
        self.atr_seq += 1
        self.after_call(ms=250, func=self.atr_end)

    def after_call(self, func=None, ms=250):
        if self.atr_stop:
            self.atr_stop = False
            self.inst_power.set_output_hp_6x74a(on_off=False)  # turn off power
            self.inst_source.set_output_agilent(on_off=False)  # turn off rf state
            self.inst_power_meter.set_rel_agilent(on_off=False)  # reference off
            pg.alert(text="ATR Stop called\n",
                     title="Stop",
                     button="확인")
        else:
            if func is not None:
                self.g_root.after(ms=ms, func=func)

    def _reset_input_offset_var(self):
        self.atr_index = 0
        self.mes_input_offset.clear()
        self.mes_output_offset.clear()

    def _reset_output_offset_var(self):
        self.atr_index = 0
        self.mes_output_offset.clear()

    def _reset_atr_var(self):
        self.excel.aging_done_var = False
        self.atr_state = "start"
        self.atr_index = 0
        self.atr_p1_var.clear()
        self.atr_input_var.clear()
        self.atr_input_curr_var.clear()
        self.atr_p_sat_var.clear()
        self.atr_p_sat_curr_var.clear()
        self.atr_overdrive_var.clear()
        self.atr_overdrive_curr_var.clear()
        self.atr_compare_0 = 0.0
        self.atr_compare_1 = 0.0
        self.atr_input_input = 0.0
        self.excel.aging_left_var = self.excel.aging_var

    def dialog_var_set(self, var_name=None, value=None):
        if (var_name is not None) and (value is not None):
            if var_name == "atr_freq":
                if value >= 1000000000000:
                    self.atr_freq.set("{0} THz".format(round(value / 1000000000000, 12)))
                elif self.excel.freq_var[0] >= 1000000000:
                    self.atr_freq.set("{0} GHz".format(round(value / 1000000000, 9)))
                elif self.excel.freq_var[0] >= 1000000:
                    self.atr_freq.set("{0} MHz".format(round(value / 1000000, 6)))
                elif self.excel.freq_var[0] >= 1000:
                    self.atr_freq.set("{0} KHz".format(round(value / 1000, 3)))
                else:
                    self.atr_freq.set("{0} Hz".format(int(value)))
            elif var_name == "atr_input_offset":
                self.atr_input_offset.set("{0} dB".format(round(value, 1)))
            elif var_name == "atr_output_offset":
                self.atr_output_offset.set("{0} dB".format(round(value, 2)))
            elif var_name == "atr_rf_input":
                self.atr_rf_input.set("{0} dBm".format(round(value, 1)))
            elif var_name == "atr_rf_output":
                self.atr_rf_output.set("{0} dBm".format(round(value, 2)))
            elif var_name == "atr_power_voltage":
                self.atr_power_voltage.set("{0} V".format(round(value, 1)))
            elif var_name == "atr_power_current":
                self.atr_current.set("{0} A".format(round(value, 2)))
            elif var_name == "aging_time_left":
                self.atr_aging_left.set("{0} sec".format(int(value)))
            else:
                print("{0} is unknown dialog var_name".format(var_name))

    @staticmethod
    def thread_sleep(sec=0.0):
        time.sleep(sec)

    def atr_stop_clicked(self):
        self.atr_stop = True

    @staticmethod
    def _atr_display_for_loop(var):
        text = ""
        for var in var:
            text += str(var) + " "
        text += "\n"
        return text

    def atr_display_clicked(self):
        names = ["frequency", "p1", "input dBm", "input watt", "input current",
                 "p_sat dBm", "p_sat watt", "p_sat current", "overdrive dBm", "overdrive watt",
                 "overdrive current"]
        units = ["Hz", "dBm", "dBm", "W", "A",
                 "dBm", "W", "A", "dBm", "W",
                 "A"]
        if len(self.excel.select_freq_var) != len(self.atr_p1_var):
            pg.alert(text="Nothing to display",
                     title="Display",
                     button="확인")
        else:
            text = ""
            pg.alert(text=text,
                     title="Display",
                     button="확인")

    def atr_sys_opt_clicked(self):
        if self.atr_system_option:
            self.atr_system_option = False
        else:
            self.atr_system_option = True

    def load_file_dialog(self):
        dir_path = fd.askopenfilename(parent=self.g_root,
                                      initialdir=os.getcwd(),
                                      title='Select Config Excel File')  # 대화창 open
        if len(dir_path) == 0:
            self.excel_path = "-1"
        else:
            check_excel = dir_path[len(dir_path) - 5:len(dir_path)]
            if check_excel == ".xlsx":
                self.excel_path = dir_path
            else:
                self.load_file_dialog()

    def save_file_dialog(self, _id):
        if _id == self.atr_save_button_id:
            init_name = "atr.xlsx"
        elif _id == self.atr_measure_input_offset_id:
            init_name = "input_offset.xlsx"
        elif _id == self.atr_measure_output_offset_id:
            init_name = "output_offset.xlsx"
        else:
            init_name = ".xlsx"
        dir_path = fd.asksaveasfilename(parent=self.g_root,
                                        initialdir=os.getcwd(),
                                        initialfile=init_name,
                                        title='Save Excel File',
                                        filetypes=[("excel files", "*.xlsx"), ("all", "*.*")]
                                        )  # 대화창 open
        if len(dir_path) == 0:
            self.excel_path = "-1"
        else:
            check_excel = dir_path[len(dir_path) - 5:len(dir_path)]
            if check_excel == ".xlsx":
                self.excel_path = dir_path
            else:
                self.save_file_dialog(_id=_id)

    def save_excel_loop_col_w(self, var_s, row, column):
        for var in var_s:
            self.excel.w_ws_list[0].cell(row, column, round((10 ** (var / 10)) / 1000, 1))
            column += 1

    def save_excel_loop_col(self, var_s, row, column):
        for var in var_s:
            self.excel.w_ws_list[0].cell(row, column, var)
            column += 1

    def save_excel_loop(self, var_s, column):
        index_start = 1
        for var in var_s:
            self.excel.w_ws_list[0].cell(index_start, ord(column) + 1 - ord("A"), var)
            index_start += 1

    def save_option(self, save_path, _id):
        names = []
        units = ["Hz", "dB", "dB"]
        if _id == self.atr_measure_input_offset_id:
            names = ["frequency", "input offset"]
        elif _id == self.atr_measure_output_offset_id:
            names = ["frequency", "output offset"]
        elif _id == self.atr_save_all_offset_id:
            names = ["frequency", "input offset", "output offset"]
        else:
            print("unknown save option error")

        self.excel.w_wb = Workbook()
        self.excel.w_ws_list.append(self.excel.w_wb.create_sheet("OFFSET"))
        self.excel.w_ws_list[0] = self.excel.w_wb.active

        if self.excel.save_hor is True:
            freq_row = 1
            offset_row_0 = 2
            offset_row_1 = 3

            name_column = "A"
            unit_column = "B"
            data_start_column = "C"

            self.save_excel_loop(var_s=names, column=name_column)
            self.excel.w_ws_list[0].column_dimensions[name_column].width = 15  # column set

            self.save_excel_loop(var_s=units, column=unit_column)

            __column = ord(data_start_column) + 1 - ord("A")
            self.save_excel_loop_col(var_s=self.excel.select_freq_var, column=__column, row=freq_row)

            if _id == self.atr_measure_input_offset_id:
                self.save_excel_loop_col(var_s=self.mes_input_offset, column=__column, row=offset_row_0)
            elif _id == self.atr_measure_output_offset_id:
                self.save_excel_loop_col(var_s=self.mes_output_offset, column=__column, row=offset_row_0)
            elif _id == self.atr_save_all_offset_id:
                self.save_excel_loop_col(var_s=self.mes_input_offset, column=__column, row=offset_row_0)
                self.save_excel_loop_col(var_s=self.mes_input_offset, column=__column, row=offset_row_1)

        self.excel.w_wb.save(save_path)

    def save_all_var(self, save_path):
        # test var
        print(self.excel.freq_var)
        print(self.excel.input_offset_var)
        print(self.excel.output_offset_var)
        print(self.excel.power_voltage_var)
        print(self.excel.aging_var)
        print(self.excel.source_com_var)
        print(self.excel.power_com_var)
        print(self.excel.power_meter_com_var)
        print(self.atr_p1_var)
        print(self.atr_input_var)
        print(self.atr_input_curr_var)
        print(self.atr_p_sat_var)
        print(self.atr_p_sat_curr_var)
        print(self.atr_overdrive_var)
        print(self.atr_overdrive_curr_var)

        # test
        self.atr_p1_var = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 41]
        self.atr_input_var = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 44]
        self.atr_input_curr_var = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 47]
        self.atr_p_sat_var = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 50]
        self.atr_p_sat_curr_var = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 53]
        self.atr_overdrive_var = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 56]
        self.atr_overdrive_curr_var = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 59]

        if self.excel.save_hor is True:
            freq_row = 1
            p1_row = 2
            input_row = 3
            input_watt_row = 4
            input_current_row = 5
            p_sat_row = 6
            p_sat_watt_row = 7
            p_sat_current_row = 8
            overdrive_row = 9
            overdrive_watt_row = 10
            overdrive_current_row = 11

            name_column = "A"
            unit_column = "B"
            data_start_column = "C"

            names = ["frequency", "p1", "input dBm", "input watt", "input current",
                     "p_sat dBm", "p_sat watt", "p_sat current", "overdrive dBm", "overdrive watt",
                     "overdrive current"]
            units = ["Hz", "dBm", "dBm", "W", "A",
                     "dBm", "W", "A", "dBm", "W",
                     "A"]

            self.excel.w_wb = Workbook()
            self.excel.w_ws_list.append(self.excel.w_wb.create_sheet("ATR"))

            self.excel.w_ws_list[0] = self.excel.w_wb.active

            self.save_excel_loop(var_s=names, column=name_column)
            self.excel.w_ws_list[0].column_dimensions[name_column].width = 15  # column set

            self.save_excel_loop(var_s=units, column=unit_column)

            __column = ord(data_start_column) + 1 - ord("A")
            self.save_excel_loop_col(var_s=self.excel.select_freq_var, column=__column, row=freq_row)
            self.save_excel_loop_col(var_s=self.atr_p1_var, column=__column, row=p1_row)
            self.save_excel_loop_col(var_s=self.atr_input_var, column=__column, row=input_row)
            self.save_excel_loop_col_w(var_s=self.atr_input_var, column=__column, row=input_watt_row)
            self.save_excel_loop_col(var_s=self.atr_input_curr_var, column=__column, row=input_current_row)
            self.save_excel_loop_col(var_s=self.atr_p_sat_var, column=__column, row=p_sat_row)
            self.save_excel_loop_col_w(var_s=self.atr_p_sat_var, column=__column, row=p_sat_watt_row)
            self.save_excel_loop_col(var_s=self.atr_p_sat_curr_var, column=__column, row=p_sat_current_row)
            self.save_excel_loop_col(var_s=self.atr_overdrive_var, column=__column, row=overdrive_row)
            self.save_excel_loop_col_w(var_s=self.atr_overdrive_var, column=__column, row=overdrive_watt_row)
            self.save_excel_loop_col_w(var_s=self.atr_overdrive_curr_var, column=__column, row=overdrive_current_row)
            self.excel.w_wb.save(save_path)
        else:  # save vertical atr
            pass

    def __find_offset_in_table(self, send_freq, select):
        searched_index = np.abs(np.array(send_freq) - self.excel.freq_var).argmin()
        if select == "output_offset":
            if send_freq == self.excel.freq_var[searched_index]:
                return self.excel.output_offset_var[searched_index]
            elif send_freq > self.excel.freq_var[searched_index]:
                if searched_index == len(self.excel.freq_var) - 1:  # range 안에 들어오는지 체크
                    print("frequency range over error")
                    return
                else:
                    return self.excel.output_offset_var[searched_index] \
                           + (send_freq - self.excel.freq_var[searched_index]) \
                           * (self.excel.output_offset_var[searched_index + 1]
                              - self.excel.output_offset_var[searched_index]) \
                           / (self.excel.freq_var[searched_index + 1] - self.excel.freq_var[searched_index])
            elif send_freq < self.excel.freq_var[searched_index]:
                if searched_index == 0:  # range 안에 들어오는지 체크
                    print("frequency range over error")
                    return
                else:
                    return self.excel.output_offset_var[searched_index - 1] \
                           + (send_freq - self.excel.freq_var[searched_index - 1]) \
                           * (self.excel.output_offset_var[searched_index]
                              - self.excel.output_offset_var[searched_index - 1]) \
                           / (self.excel.freq_var[searched_index] - self.excel.freq_var[searched_index - 1])
        elif select == "input_offset":
            if send_freq == self.excel.freq_var[searched_index]:
                return self.excel.input_offset_var[searched_index]
            elif send_freq > self.excel.freq_var[searched_index]:
                if searched_index == len(self.excel.freq_var) - 1:  # range 안에 들어오는지 체크
                    print("frequency range over error")
                    return
                else:
                    return self.excel.input_offset_var[searched_index] \
                           + (send_freq - self.excel.freq_var[searched_index]) \
                           * (self.excel.input_offset_var[searched_index + 1]
                              - self.excel.input_offset_var[searched_index]) \
                           / (self.excel.freq_var[searched_index + 1]
                              - self.excel.freq_var[searched_index])
            elif send_freq < self.excel.freq_var[searched_index]:
                if searched_index == 0:  # range 안에 들어오는지 체크
                    print("frequency range over error")
                    return
                else:
                    return self.excel.input_offset_var[searched_index - 1] \
                           + (send_freq - self.excel.freq_var[searched_index - 1]) \
                           * (self.excel.input_offset_var[searched_index]
                              - self.excel.input_offset_var[searched_index - 1]) \
                           / (self.excel.freq_var[searched_index]
                              - self.excel.freq_var[searched_index - 1])
            else:
                print("error input, output offset call")

    def __atr_inst_call(self):
        self.inst_source = Instrument.Source()
        self.inst_power = Instrument.PowerSupply()
        self.inst_power_meter = Instrument.PowerMeter()
        # source open
        if self.excel.source_comm_option == "GPIB":
            self.inst_source.gpib_address = self.excel.source_com_var
            self.inst_source.open_instrument_gpib(gpib_address=self.inst_source.gpib_address)
        elif self.excel.source_comm_option == "USB":
            pass
        elif self.excel.source_comm_option == "SERIAL":
            pass
        # power supply open
        if self.excel.power_comm_option == "GPIB":
            self.inst_power.gpib_address = self.excel.power_com_var
            # test for power except
            self.inst_power.open_instrument_gpib(gpib_address=self.inst_power.gpib_address)
        elif self.excel.power_comm_option == "USB":
            pass
        elif self.excel.power_comm_option == "SERIAL":
            pass
        # power meter open
        if self.excel.power_meter_comm_option == "GPIB":
            self.inst_power_meter.gpib_address = self.excel.power_meter_com_var
            self.inst_power_meter.open_instrument_gpib(gpib_address=self.inst_power_meter.gpib_address)
        elif self.excel.power_meter_comm_option == "USB":
            pass
        elif self.excel.power_meter_comm_option == "SERIAL":
            pass

    def __atr_get_idn(self):
        print(self.inst_source.query_instrument("*IDN?"))
        # test for power except
        print(self.inst_power.query_instrument("*IDN?"))
        print(self.inst_power_meter.query_instrument("*IDN?"))

    def atr_error_state(self, text):
        self.atr_state = ""
        print("p1 procedure error")
        pg.alert(text=text + "\n",
                 title="Error",
                 button="확인")
